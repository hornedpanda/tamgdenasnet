from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Cost

path = 'tamgdenasnet/raw_data/cost2023.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['cost']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        row = dataframe1.max_row
        column = dataframe1.max_column
        # dataframe1 = pd.read_excel('raw_data/GPI-2023-overall-scores-and-domains-2008-2023.xlsx')
        # print(dataframe1)
        # for row in range(4, dataframe1.max_row):
            # for col in dataframe1.iter_cols(1, dataframe1.max_column):
            # print(col['004'].value)
        print("\nValue of first column")
        for i in range(1, row + 1):
            # k = i - 4
            rating = dataframe1.cell(row=i, column=1)
            country = dataframe1.cell(row=i, column=2)
            cost_index = dataframe1.cell(row=i, column=3)
            rent_index = dataframe1.cell(row=i, column=4)
            cost_rent_index = dataframe1.cell(row=i, column=5)
            salary = dataframe1.cell(row=i, column=8)

            # print('У страны {} место {}, salary {}'.format(country.value, str(rating.value), str(salary.value)))
            country, created = Country.objects.update_or_create(
                name=country.value
            )
            cost, created = Cost.objects.update_or_create(
                country=country,
                year=2023,
                defaults={'rating': rating.value,
                          'cost_index': cost_index.value,
                          'rent_index': rent_index.value,
                          'cost_rent_index': cost_rent_index.value,
                          'salary': salary.value,
                          'year': 2023,
                          },
                )
            print(cost)
