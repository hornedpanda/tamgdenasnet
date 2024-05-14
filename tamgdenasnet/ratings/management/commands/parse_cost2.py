from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Rating, Country_Rating

path = 'tamgdenasnet/raw_data/cost2023.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['cost']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        row = dataframe1.max_row
        column = dataframe1.max_column
        year = 2023
        # dataframe1 = pd.read_excel('raw_data/GPI-2023-overall-scores-and-domains-2008-2023.xlsx')
        # print(dataframe1)
        # for row in range(4, dataframe1.max_row):
            # for col in dataframe1.iter_cols(1, dataframe1.max_column):
            # print(col['004'].value)
        # cost, rent, salary = 0, 0, 0
        rating_values = {'Numbeo Cost of Living Index': 0,
                        'Numbeo Rent Index': 0,
                        'Numbeo Local Purchasing Power': 0}
        for name in rating_values:
            rating, created = Rating.objects.update_or_create(
                    # name='Global Peace Index GPI',
                    name=name,
                    # defaults={
                    #     'name': 'Global Peace Index GPI',
                    #     },
                )
            if created:
                print(f'created new rating {rating.name}')
        print("\nValue of first column")
        for i in range(1, row + 1):
            # k = i - 4
            # rating = dataframe1.cell(row=i, column=1)
            country = dataframe1.cell(row=i, column=2)
            rating_values['Numbeo Cost of Living Index'] = dataframe1.cell(row=i, column=3).value
            rating_values['Numbeo Rent Index'] = dataframe1.cell(row=i, column=4).value
                # cost_rent_index = dataframe1.cell(row=i, column=5)
            rating_values['Numbeo Local Purchasing Power'] = dataframe1.cell(row=i, column=8).value

            # print('У страны {} cost {}, rent {}, salary {}'.format(country.value,
            #       str(cost.value), str(rent.value),
            #       str(salary.value)))
            country, created = Country.objects.update_or_create(
                name=country.value
            )
            if created:
                print(f'created new country {country}')
            for rating in Rating.objects.filter(name__in=rating_values):
                print(rating.name)
                # rating_value = rating_values[rating.name]
                country_rating, created = Country_Rating.objects.update_or_create(
                    country=country,
                    rating=rating,
                    year=year,
                    defaults={'value': rating_values[rating.name],
                            # 'place': k,
                            #   'year': 2023,
                            },
                    )
                print(country_rating)



            # cost, created = Cost.objects.update_or_create(
            #     country=country,
            #     year=2023,
            #     defaults={'rating': rating.value,
            #               'cost_index': cost_index.value,
            #               'rent_index': rent_index.value,
            #               'cost_rent_index': cost_rent_index.value,
            #               'salary': salary.value,
            #               'year': 2023,
            #               },
            #     )
            # print(cost)
