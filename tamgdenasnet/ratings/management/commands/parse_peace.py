from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Peace

path = 'tamgdenasnet/raw_data/GPI-2023.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['Overall Scores']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        row = dataframe1.max_row
        column = dataframe1.max_column
        # dataframe1 = pd.read_excel('raw_data/GPI-2023-overall-scores-and-domains-2008-2023.xlsx')
        # print(dataframe1)
        # for row in range(4, dataframe1.max_row):
            # for col in dataframe1.iter_cols(1, dataframe1.max_column):
            # print(col['004'].value)
        print("\nValue of first column")
        for i in range(5, row + 1):
            k = i - 4
            cell_obj1 = dataframe1.cell(row=i, column=1)
            cell_obj2 = dataframe1.cell(row=i, column=18)
            print('У страны {} рейтинг {}'.format(cell_obj1.value, str(cell_obj2.value)))
            country, created = Country.objects.update_or_create(
                name=cell_obj1.value
            )
            peace, created = Peace.objects.update_or_create(
                country=country,
                year=2023,
                defaults={'value': cell_obj2.value,
                          'rating': k,
                          'year': 2023,
                          },
                )
            print(peace)
