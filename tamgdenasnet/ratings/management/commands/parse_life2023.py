from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Rating, Country_Rating

path = 'tamgdenasnet/raw_data/life2023.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['life']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        row = dataframe1.max_row
        column = dataframe1.max_column
        year = 2023
        name = "Life Expectancy"
        rating, created = Rating.objects.update_or_create(
                    # name='Global Peace Index GPI',
                    name=name,
                    defaults={
                        'decreasing': 0,
                        },
                )
        if created:
            print(f'Created new rating {rating.name}')
        for i in range(2, row + 1):
            country = dataframe1.cell(row=i, column=2)
            life_exp = dataframe1.cell(row=i, column=3).value
            print('У гражданина {} срок годности {} лет'.format(country.value,
                  str(life_exp)))
            country, created = Country.objects.update_or_create(
                name=country.value
            )
            if created:
                print(f'Created new country {country}')
            country_rating, created = Country_Rating.objects.update_or_create(
                    country=country,
                    rating=rating,
                    year=year,
                    defaults={'value': life_exp,
                              # 'place': k,
                              #   'year': 2023,
                              },
                    )
            print(country_rating)