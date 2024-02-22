from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Climate

path = 'tamgdenasnet/raw_data/quality2023.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['quality']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        row = dataframe1.max_row
        column = dataframe1.max_column
        print("\nValue of first column")
        cli_rating = []
        for i in range(1, row + 1):
            # k = i - 4
            # rating = dataframe1.cell(row=i, column=1)
            country = dataframe1.cell(row=i, column=2).value
            value = dataframe1.cell(row=i, column=11).value
            cli_rating.append([country, value])
            # print('У страны {} индекс {}'.format(country.value, str(value.value)))
            

            # climate, created = Climate.objects.update_or_create(
            #     country=country,
            #     year=2023,
            #     defaults={'rating': rating.value,
            #               'value': value.value,
            #               'year': 2023,
            #               },
            #     )
            # print(country, value)
        cli_rating = sorted(cli_rating, key=lambda country: -country[1])
        i = 1
        for value in cli_rating:
            # print(country)
            country, created = Country.objects.update_or_create(
                name=value[0]
            )
            # print(i, country[0], country[1])
            climate, created = Climate.objects.update_or_create(
                country=country,
                year=2023,
                defaults={'rating': i,
                          'value': value[1],
                          'year': 2023,
                          },
                )
            i += 1
        print(Climate.objects.all())