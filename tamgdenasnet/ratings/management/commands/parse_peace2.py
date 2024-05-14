from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Rating, Country_Rating

path = 'tamgdenasnet/raw_data/GPI-2022.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['Overall Scores']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        row = dataframe1.max_row
        # column = dataframe1.max_column
        print("\nValue of first column")
        for i in range(5, row + 1):
            k = i - 4
            cell_obj1 = dataframe1.cell(row=i, column=1)
            cell_obj2 = dataframe1.cell(row=i, column=17)
            # print('У страны {} рейтинг {}'.format(cell_obj1.value, str(cell_obj2.value)))
            country, created = Country.objects.update_or_create(
                name=cell_obj1.value,
                # slug=cell_obj1.value
                # defaults={
                #     'name': (cell_obj1.value),
                #     },
            )
            if created:
                print(f'created new country {cell_obj1.value}')
            rating, created = Rating.objects.update_or_create(
                # name='Global Peace Index GPI',
                name='Global Peace Index GPI',
                # defaults={
                #     'name': 'Global Peace Index GPI',
                #     },
            )
            if created:
                print(f'created new rating {rating.name}')
            country_rating, created = Country_Rating.objects.update_or_create(
                country=country,
                rating=rating,
                year=2022,
                defaults={'value': cell_obj2.value,
                        #   'place': k,
                        #   'year': 2023,
                          },
                )
            print(country_rating)
