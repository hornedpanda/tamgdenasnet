from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Rating, Country_Rating

path = 'tamgdenasnet/raw_data/Temperature2023.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['temperature']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        row = dataframe1.max_row
        column = dataframe1.max_column
        year = 2022
        # dataframe1 = pd.read_excel('raw_data/GPI-2023-overall-scores-and-domains-2008-2023.xlsx')
        # print(dataframe1)
        # for row in range(4, dataframe1.max_row):
            # for col in dataframe1.iter_cols(1, dataframe1.max_column):
            # print(col['004'].value)
        # cost, rent, salary = 0, 0, 0
        name = "Average Yearly Temperature"
        rating, created = Rating.objects.update_or_create(
                    # name='Global Peace Index GPI',
                    name=name,
                    defaults={
                        'decreasing': 0,
                        },
                )
        if created:
            print(f'Created new rating {rating.name}')
        for i in range(1, row + 1):
            country = dataframe1.cell(row=i, column=1)
            temperature = dataframe1.cell(row=i, column=2).value
            # print('У страны {} температура {}'.format(country.value,
            #       str(temperature)))
            country, created = Country.objects.update_or_create(
                name=country.value
            )
            if created:
                print(f'Created new country {country}')
            country_rating, created = Country_Rating.objects.update_or_create(
                    country=country,
                    rating=rating,
                    year=year,
                    defaults={'value': temperature,
                            # 'place': k,
                            #   'year': 2023,
                            },
                    )
            print(country_rating)