from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Peace, Democracy

path = 'tamgdenasnet/raw_data/democracy.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['democracy-index-eiu']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        max_row = dataframe1.max_row
        column = dataframe1.max_column
        for row in range(2, max_row + 1):
            country = dataframe1.cell(row=row, column=1).value
            year = dataframe1.cell(row=row, column=3).value
            if year < 2022:
                continue
            value = dataframe1.cell(row=row, column=4).value
            value = round(float(value), 2)
            # print(f"{country} имеет рейтинг {value} в Democracy Index в {year} году")
            country, created = Country.objects.update_or_create(
                name=country
            )
            if created:
                print('new country ', country)
            democracy, created = Democracy.objects.update_or_create(
                country=country,
                year=2022,
                defaults={'value': value,
                          'year': 2022,
                          },
                )
            if created:
                print('new record ', democracy)
        democracy = Democracy.objects.order_by('-value')
        rating = 1
        for country in democracy:
            country.rating = rating
            country.save()
            print(country)
            rating += 1
