from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Quality

path = 'tamgdenasnet/raw_data/quality2023.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['quality']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        row = dataframe1.max_row
        column = dataframe1.max_column
        print("\nValue of first column")
        for i in range(1, row + 1):
            # k = i - 4
            rating = dataframe1.cell(row=i, column=1)
            country = dataframe1.cell(row=i, column=2)
            value = dataframe1.cell(row=i, column=3)

            # print('У страны {} место {}, индекс {}'.format(country.value, str(rating.value), str(value.value)))
            country, created = Country.objects.update_or_create(
                name=country.value
            )
            quality, created = Quality.objects.update_or_create(
                country=country,
                year=2023,
                defaults={'rating': rating.value,
                          'value': value.value,
                          'year': 2023,
                          },
                )
            print(quality)