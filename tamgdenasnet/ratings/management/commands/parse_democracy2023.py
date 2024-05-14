from django.core.management import BaseCommand
import openpyxl, os

from ratings.models import Country, Rating, Country_Rating

path = 'tamgdenasnet/raw_data/democracy2023.xlsx'
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe['Лист1']


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('hello')
        max_row = dataframe1.max_row
        column = dataframe1.max_column
        k = 1
        for row in range(1, max_row + 1):
            country = dataframe1.cell(row=row, column=3).value[1:]
            year = 2023
            value = dataframe1.cell(row=row, column=5).value
            value = round(float(value), 2)
            # print(f"{country} имеет рейтинг {value} в Democracy Index в {year} году")
            country, created = Country.objects.update_or_create(
                name=country
            )
            # if created:
            #     print('new country ', country)
            if created:
                print(f'created new country {country}')
            rating, created = Rating.objects.update_or_create(
                # name='Global Peace Index GPI',
                name='The Economist Democracy Index',
                # defaults={
                #     'name': 'Global Peace Index GPI',
                #     },
            )
            if created:
                print(f'created new rating {rating.name}')
            country_rating, created = Country_Rating.objects.update_or_create(
                country=country,
                rating=rating,
                year=year,
                defaults={'value': value,
                          #   'place': k,
                          #   'year': 2023,
                          },
                )
            print(country_rating)
        #     k += 1
        # country_rating = Country_Rating.objects.filter(
        #     rating__name='The Economist Democracy Index').filter(
        #         year=year).order_by('-value')
        # print(country_rating)
        # place = 1
        # for country in country_rating:
        #     country.place = place
        #     country.save()
        #     print(country)
        #     place += 1
